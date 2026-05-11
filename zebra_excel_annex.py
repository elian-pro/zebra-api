"""
zebra_excel_annex.py

Genera el Excel anexo de Proyección de Inversión para acompañar la propuesta DOCX.
Replica la estructura de las dos hojas del DEMO original:
  - Hoja 1: Calculadora de Inversión (con inputs editables y fórmulas)
  - Hoja 2: Plan de Inversión — Proyección a 12 meses

Uso:
    from zebra_excel_annex import generate_investment_excel
    generate_investment_excel(calc_output, plan_output, output_path)
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule


# Paleta Zebra
YELLOW = "F9D626"
BLACK = "111111"
DARK_GREY = "333333"
MID_GREY = "888888"
LIGHT_GREY = "DDDDDD"
BG_LIGHT = "F5F5F5"
WHITE = "FFFFFF"


def _border(color=LIGHT_GREY, style="thin"):
    side = Side(border_style=style, color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _bottom_border(color=LIGHT_GREY, style="thin"):
    side = Side(border_style=style, color=color)
    return Border(bottom=side)


def _fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _label_cell(cell, text, bold=False, size=10, color=DARK_GREY, fill_color=None):
    cell.value = text
    cell.font = Font(name="Arial", size=size, bold=bold, color=color)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    if fill_color:
        cell.fill = _fill(fill_color)


def _value_cell(cell, value, bold=False, size=10, color=DARK_GREY,
                fill_color=None, fmt=None, align="right"):
    cell.value = value
    cell.font = Font(name="Arial", size=size, bold=bold, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if fill_color:
        cell.fill = _fill(fill_color)
    if fmt:
        cell.number_format = fmt


def _header_cell(cell, text, fill_color=BLACK, font_color=YELLOW, size=10):
    cell.value = text
    cell.font = Font(name="Arial", size=size, bold=True, color=font_color)
    cell.fill = _fill(fill_color)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _build_calc_sheet(ws, calc_output):
    """
    Hoja 1 — Calculadora de Inversión.
    Reproduce la lógica del DEMO con valores del cliente.
    Los inputs editables están marcados con celdas color amarillo claro.
    """
    from zebra_investment_calc import (
        LEADS_DIA_POR_ASESOR, DIAS_PROMEDIO_MES,
        VENTANA_CITA_DIAS, VENTANA_APARTADO_DIAS, VENTANA_CIERRE_DIAS,
    )

    ws.title = "Calculadora de Inversion"

    # Anchos de columna
    widths = {"A": 2, "B": 32, "C": 16, "D": 18, "E": 2, "F": 24, "G": 16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Título principal
    ws.merge_cells("B2:D3")
    title = ws["B2"]
    _header_cell(title, "CALCULADORA DE INVERSIÓN", fill_color=BLACK, font_color=YELLOW, size=14)

    # Subtítulo
    ws.merge_cells("B4:D4")
    sub = ws["B4"]
    _label_cell(sub, "Ingeniería inversa: Valor de Proyecto y Tiempo de Absorción",
                size=9, color=MID_GREY)
    sub.alignment = Alignment(horizontal="center")

    # ------ INPUTS ------
    ws["B6"] = "INPUTS DEL MODELO"
    ws["B6"].font = Font(name="Arial", size=9, bold=True, color=MID_GREY)

    # Inputs editables (fondo amarillo claro)
    input_rows = [
        ("Unidades Totales", calc_output.unidades_totales, "0", 7),
        ("Ticket Promedio", calc_output.ticket_promedio, '"$"#,##0" MXN"', 8),
        ("Absorción del Proyecto (meses)", calc_output.absorcion_meses, "0", 9),
        ("CPL Estimado", calc_output.cpl, '"$"#,##0" MXN"', 10),
        ("% Inversión sobre Valor", calc_output.porcentaje_inversion, "0.00%", 11),
    ]
    for label, value, fmt, row in input_rows:
        _label_cell(ws[f"B{row}"], label, size=10)
        _value_cell(ws[f"D{row}"], value, bold=True, fmt=fmt, fill_color="FFF8C9")
        ws[f"D{row}"].border = _border(color=YELLOW)

    # ------ TASAS DE CONVERSIÓN (editables) ------
    ws["B13"] = "TASAS DE CONVERSIÓN"
    ws["B13"].font = Font(name="Arial", size=9, bold=True, color=MID_GREY)

    tasa_rows = [
        ("% Conversión a Cita", calc_output.tasas['cita'], 14),
        ("% Asistencia", calc_output.tasas['asistencia'], 15),
        ("% Apartado", calc_output.tasas['apartado'], 16),
        ("% Cierre", calc_output.tasas['cierre'], 17),
    ]
    for label, value, row in tasa_rows:
        _label_cell(ws[f"B{row}"], label, size=10)
        _value_cell(ws[f"D{row}"], value, bold=True, fmt="0.0%", fill_color="FFF8C9")
        ws[f"D{row}"].border = _border(color=YELLOW)

    # ------ OUTPUTS CALCULADOS (con fórmulas vinculadas) ------
    ws["B19"] = "RESULTADOS CALCULADOS"
    ws["B19"].font = Font(name="Arial", size=9, bold=True, color=MID_GREY)

    # Valor del Proyecto = Unidades × Ticket
    _label_cell(ws["B20"], "Valor del Proyecto", size=10)
    _value_cell(ws["D20"], "=D7*D8", bold=True, fmt='"$"#,##0" MXN"', fill_color=BG_LIGHT)

    # CPA Estimado = % Inversión × Ticket
    _label_cell(ws["B21"], "CPA Estimado", size=10)
    _value_cell(ws["D21"], "=D11*D8", bold=True, fmt='"$"#,##0" MXN"', fill_color=BG_LIGHT)

    # Presupuesto Total = Valor × % Inversión
    _label_cell(ws["B22"], "Presupuesto Total", size=10)
    _value_cell(ws["D22"], "=D20*D11", bold=True, fmt='"$"#,##0" MXN"', fill_color=BG_LIGHT)

    # ROA = Valor / Presupuesto
    _label_cell(ws["B23"], "ROA Esperado", size=10)
    _value_cell(ws["D23"], "=D20/D22", bold=True, fmt='0.0"x"', fill_color=BG_LIGHT)

    # Unidades objetivo mensuales
    _label_cell(ws["B24"], "Unidades Objetivo Mensuales", size=10)
    _value_cell(ws["D24"], "=D7/D9", bold=True, fmt="0.00", fill_color=BG_LIGHT)

    # Inversión en Pauta Mensual (destacada)
    _label_cell(ws["B25"], "INVERSIÓN PAUTA MENSUAL", bold=True, size=11, fill_color=YELLOW)
    _value_cell(ws["D25"], "=D22/D9", bold=True, fmt='"$"#,##0" MXN"',
                fill_color=YELLOW, color=BLACK, size=11)

    # Leads esperados mensuales
    _label_cell(ws["B26"], "Leads Esperados (mensuales)", size=10)
    _value_cell(ws["D26"], "=ROUNDDOWN(D25/D10,0)", bold=True, fmt="#,##0", fill_color=BG_LIGHT)

    # Capacidad max por asesor
    _label_cell(ws["B27"], "Capacidad Máx. Mensual / Asesor", size=10)
    _value_cell(ws["D27"], f"=ROUNDDOWN({LEADS_DIA_POR_ASESOR}*{DIAS_PROMEDIO_MES},0)",
                fmt="0", fill_color=BG_LIGHT)

    # Sala necesaria
    _label_cell(ws["B28"], "Sala Necesaria (asesores)", bold=True, size=10)
    _value_cell(ws["D28"], "=ROUNDDOWN(D26/D27,0)", bold=True, fmt="0", fill_color=BG_LIGHT)

    # ------ CASCADA DEL FUNNEL ------
    ws["F6"] = "EFICIENCIA COMERCIAL ESPERADA"
    ws["F6"].font = Font(name="Arial", size=9, bold=True, color=MID_GREY)

    funnel_rows = [
        ("Citas (mes)", "=D26*D14", 8),
        ("Asistencias (mes)", "=G8*D15", 9),
        ("Apartados (mes)", "=G9*D16", 10),
        ("Cierres (mes)", "=G10*D17", 11),
    ]
    for label, formula, row in funnel_rows:
        _label_cell(ws[f"F{row}"], label, size=10)
        _value_cell(ws[f"G{row}"], formula, bold=True, fmt="0", fill_color=BG_LIGHT)

    # ------ NOTAS ------
    ws.merge_cells("B30:G31")
    notas = ws["B30"]
    _label_cell(
        notas,
        "Notas: El % de inversión sobre valor del proyecto es un promedio fijo (3%) "
        "como punto de partida; se ajusta según ticket, ubicación y dinámica del proyecto. "
        "CPL y tasas de conversión basadas en benchmarks Zebra Real Estate; pueden "
        "ajustarse al rendimiento real del cliente.",
        size=8, color=MID_GREY
    )
    notas.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Altura de filas
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22
    for r in range(6, 32):
        ws.row_dimensions[r].height = 20


def _build_plan_sheet(ws, calc_output, plan_output):
    """
    Hoja 2 — Plan de Inversión 12 meses.
    Replica la estructura del DEMO con datos del cliente.
    """
    from zebra_investment_calc import (
        LEADS_DIA_POR_ASESOR, DIAS_PROMEDIO_MES,
        VENTANA_CITA_DIAS, VENTANA_APARTADO_DIAS, VENTANA_CIERRE_DIAS,
        MULTIPLICADORES_UNIDADES,
    )

    ws.title = "Plan 12 Meses"

    # Anchos
    ws.column_dimensions["A"].width = 36
    for c in range(2, 19):
        ws.column_dimensions[get_column_letter(c)].width = 12

    # Título
    ws.merge_cells("A1:R1")
    title = ws["A1"]
    _header_cell(title, "PLAN DE INVERSIÓN — PROYECCIÓN A 12 MESES",
                  fill_color=BLACK, font_color=YELLOW, size=14)

    ws.merge_cells("A2:R2")
    sub = ws["A2"]
    _label_cell(sub,
                "Zebra High Performance Marketing | Implementación + Operación | "
                "6 ventanas de conversión por mes",
                size=9, color=MID_GREY)
    sub.alignment = Alignment(horizontal="center")

    # ------ SUPUESTOS ------
    ws["A4"] = "SUPUESTOS DEL MODELO"
    ws["A4"].font = Font(name="Arial", size=9, bold=True, color=MID_GREY)

    supuestos = [
        ("CPL Proyectado", calc_output.cpl, '"$"#,##0" MXN"'),
        ("Leads / día / asesor", LEADS_DIA_POR_ASESOR, "0"),
        ("Días promedio del mes", DIAS_PROMEDIO_MES, "0.0"),
        ("% Conversión a Cita", calc_output.tasas['cita'], "0.0%"),
        ("% Asistencia", calc_output.tasas['asistencia'], "0.0%"),
        ("% Apartado", calc_output.tasas['apartado'], "0.0%"),
        ("% Cierre", calc_output.tasas['cierre'], "0.0%"),
        ("Ventana Cita (días)", VENTANA_CITA_DIAS, "0"),
        ("Ventana Apartado (días)", VENTANA_APARTADO_DIAS, "0"),
        ("Ventana Cierre (días)", VENTANA_CIERRE_DIAS, "0"),
        ("Precio por propiedad base", calc_output.ticket_promedio, '"$"#,##0" MXN"'),
    ]
    row = 5
    for label, value, fmt in supuestos:
        _label_cell(ws.cell(row=row, column=1), label, size=9)
        _value_cell(ws.cell(row=row, column=2), value, bold=True, fmt=fmt, fill_color="FFF8C9")
        row += 1

    # ------ PROYECCIÓN MENSUAL ------
    start_row = 18
    ws.cell(row=start_row, column=1).value = "PROYECCIÓN MENSUAL — 12 MESES / 4 TRIMESTRES"
    ws.cell(row=start_row, column=1).font = Font(
        name="Arial", size=10, bold=True, color=BLACK
    )

    # Header row
    hr = start_row + 1
    headers = ["CONCEPTO"] + [f"M{i}" for i in range(1, 4)] + ["Q1"] \
              + [f"M{i}" for i in range(4, 7)] + ["Q2"] \
              + [f"M{i}" for i in range(7, 10)] + ["Q3"] \
              + [f"M{i}" for i in range(10, 13)] + ["Q4"] \
              + ["TOTAL"]
    for i, h in enumerate(headers, start=1):
        _header_cell(ws.cell(row=hr, column=i), h, size=9)

    def _is_quarter_col(col_idx):
        # cols 5, 9, 13, 17 son trimestres; col 18 es total
        return col_idx in [5, 9, 13, 17, 18]

    # Equipo e inversión
    cur = hr + 2
    ws.cell(row=cur, column=1).value = "EQUIPO E INVERSIÓN"
    ws.cell(row=cur, column=1).font = Font(name="Arial", size=9, bold=True, color=MID_GREY)
    cur += 1

    # Asesores activos
    _label_cell(ws.cell(row=cur, column=1), "Asesores activos", size=9)
    for i, m in enumerate(plan_output.meses):
        col = _month_to_col(i + 1)
        _value_cell(ws.cell(row=cur, column=col), m.asesores_activos,
                    fmt="0", align="center")
    # Trimestres y total
    for q_col, q in zip([5, 9, 13, 17], plan_output.trimestres):
        _value_cell(ws.cell(row=cur, column=q_col), q.asesores_promedio,
                    bold=True, fmt="0.0", align="center", fill_color=BG_LIGHT)
    _value_cell(ws.cell(row=cur, column=18), plan_output.asesores_promedio_anual,
                bold=True, fmt="0.0", align="center", fill_color=BG_LIGHT)
    cur += 1

    # Leads
    _label_cell(ws.cell(row=cur, column=1), "Leads generados", size=9)
    for i, m in enumerate(plan_output.meses):
        col = _month_to_col(i + 1)
        _value_cell(ws.cell(row=cur, column=col), int(m.leads_generados),
                    fmt="#,##0", align="center")
    for q_col, q in zip([5, 9, 13, 17], plan_output.trimestres):
        _value_cell(ws.cell(row=cur, column=q_col), int(q.leads_total),
                    bold=True, fmt="#,##0", align="center", fill_color=BG_LIGHT)
    _value_cell(ws.cell(row=cur, column=18), int(plan_output.leads_total_anual),
                bold=True, fmt="#,##0", align="center", fill_color=BG_LIGHT)
    cur += 1

    # Inversión pauta
    _label_cell(ws.cell(row=cur, column=1), "Inversión en pauta", size=9, bold=True)
    for i, m in enumerate(plan_output.meses):
        col = _month_to_col(i + 1)
        _value_cell(ws.cell(row=cur, column=col), m.inversion_pauta,
                    fmt='"$"#,##0', align="center")
    for q_col, q in zip([5, 9, 13, 17], plan_output.trimestres):
        _value_cell(ws.cell(row=cur, column=q_col), q.inversion_total,
                    bold=True, fmt='"$"#,##0', align="center", fill_color=YELLOW, color=BLACK)
    _value_cell(ws.cell(row=cur, column=18), plan_output.inversion_total_anual,
                bold=True, fmt='"$"#,##0', align="center", fill_color=YELLOW, color=BLACK)
    cur += 2

    # ------ EMBUDO REAL ------
    ws.cell(row=cur, column=1).value = "EMBUDO REAL (con ventanas de conversión)"
    ws.cell(row=cur, column=1).font = Font(name="Arial", size=9, bold=True, color=MID_GREY)
    cur += 1

    funnel_rows = [
        ("Citas reales", "citas_reales", "0"),
        ("Asistencias reales", "asistencias_reales", "0"),
        ("Apartados reales", "apartados_reales", "0"),
        ("Cierres reales", "cierres_reales", "0"),
    ]
    for label, attr, fmt in funnel_rows:
        is_cierres = (attr == "cierres_reales")
        _label_cell(ws.cell(row=cur, column=1), label, size=9, bold=is_cierres)
        for i, m in enumerate(plan_output.meses):
            col = _month_to_col(i + 1)
            val = getattr(m, attr)
            _value_cell(ws.cell(row=cur, column=col), val,
                        fmt=fmt, align="center")
        for q_col, q in zip([5, 9, 13, 17], plan_output.trimestres):
            val = getattr(q, attr)
            _value_cell(ws.cell(row=cur, column=q_col), val,
                        bold=True, fmt=fmt, align="center", fill_color=BG_LIGHT)
        # Total anual (suma de meses)
        total = sum(getattr(m, attr) for m in plan_output.meses)
        _value_cell(ws.cell(row=cur, column=18), total,
                    bold=True, fmt=fmt, align="center", fill_color=BG_LIGHT)
        cur += 1

    cur += 1

    # ------ INGRESOS POR ESCENARIO ------
    ws.cell(row=cur, column=1).value = "PROYECCIÓN DE INGRESOS (Cierres × Ticket × Multiplicador)"
    ws.cell(row=cur, column=1).font = Font(name="Arial", size=9, bold=True, color=MID_GREY)
    cur += 1

    for mult in MULTIPLICADORES_UNIDADES:
        _label_cell(ws.cell(row=cur, column=1),
                     f"Ingresos @ {mult:.1f} unidad/operación", size=9)
        for i, m in enumerate(plan_output.meses):
            col = _month_to_col(i + 1)
            _value_cell(ws.cell(row=cur, column=col), m.ingresos_por_escenario[mult],
                        fmt='"$"#,##0', align="center")
        for q_col, q in zip([5, 9, 13, 17], plan_output.trimestres):
            _value_cell(ws.cell(row=cur, column=q_col), q.ingresos_por_escenario[mult],
                        bold=True, fmt='"$"#,##0', align="center", fill_color=BG_LIGHT)
        _value_cell(ws.cell(row=cur, column=18),
                    plan_output.ingresos_anual_por_escenario[mult],
                    bold=True, fmt='"$"#,##0', align="center", fill_color=BG_LIGHT)
        cur += 1

    cur += 1

    # ------ MARGEN ------
    ws.cell(row=cur, column=1).value = "MARGEN DE ADQUISICIÓN (Ingresos − Inversión)"
    ws.cell(row=cur, column=1).font = Font(name="Arial", size=9, bold=True, color=MID_GREY)
    cur += 1

    for mult in MULTIPLICADORES_UNIDADES:
        _label_cell(ws.cell(row=cur, column=1),
                     f"Margen @ {mult:.1f} unidad/operación", size=9)
        for i, m in enumerate(plan_output.meses):
            col = _month_to_col(i + 1)
            _value_cell(ws.cell(row=cur, column=col), m.margen_por_escenario[mult],
                        fmt='"$"#,##0', align="center")
        for q_col, q in zip([5, 9, 13, 17], plan_output.trimestres):
            _value_cell(ws.cell(row=cur, column=q_col), q.margen_por_escenario[mult],
                        bold=True, fmt='"$"#,##0', align="center", fill_color=BG_LIGHT)
        _value_cell(ws.cell(row=cur, column=18),
                    plan_output.margen_anual_por_escenario[mult],
                    bold=True, fmt='"$"#,##0', align="center", fill_color=BG_LIGHT)
        cur += 1

    cur += 1

    # ------ ROA ACUMULADO ------
    ws.cell(row=cur, column=1).value = "ROA ACUMULADO (Ingresos / Inversión)"
    ws.cell(row=cur, column=1).font = Font(name="Arial", size=9, bold=True, color=MID_GREY)
    cur += 1

    for mult in MULTIPLICADORES_UNIDADES:
        _label_cell(ws.cell(row=cur, column=1),
                     f"ROA @ {mult:.1f} unidad/operación", size=9)
        for i, m in enumerate(plan_output.meses):
            col = _month_to_col(i + 1)
            _value_cell(ws.cell(row=cur, column=col), m.roa_por_escenario[mult],
                        fmt='0.0"x"', align="center")
        for q_col, q in zip([5, 9, 13, 17], plan_output.trimestres):
            _value_cell(ws.cell(row=cur, column=q_col), q.roa_por_escenario[mult],
                        bold=True, fmt='0.0"x"', align="center", fill_color=BG_LIGHT)
        _value_cell(ws.cell(row=cur, column=18),
                    plan_output.roa_anual_por_escenario[mult],
                    bold=True, fmt='0.0"x"', align="center", fill_color=BG_LIGHT)
        cur += 1


def _month_to_col(month):
    """Mapea mes (1-12) a columna del Excel:
    M1=2, M2=3, M3=4, Q1=5, M4=6, M5=7, M6=8, Q2=9, ...
    """
    if month <= 3:
        return month + 1  # M1=2, M2=3, M3=4
    elif month <= 6:
        return month + 2  # M4=6, M5=7, M6=8
    elif month <= 9:
        return month + 3  # M7=10, M8=11, M9=12
    else:
        return month + 4  # M10=14, M11=15, M12=16


def generate_investment_excel(calc_output, plan_output, output_path):
    """
    Genera el Excel anexo con dos hojas:
      - Calculadora de Inversión (con fórmulas vinculadas, inputs editables)
      - Plan de Inversión 12 meses (con valores calculados)
    """
    wb = Workbook()
    ws1 = wb.active
    _build_calc_sheet(ws1, calc_output)

    ws2 = wb.create_sheet("Plan 12 Meses")
    _build_plan_sheet(ws2, calc_output, plan_output)

    wb.save(output_path)
    return output_path
